from __future__ import annotations

import csv
import json
from io import BytesIO, TextIOWrapper
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence
from zipfile import ZipFile



def read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return [row for row in reader]


def normalize_header(value: str) -> str:
    return "".join(ch.lower() for ch in value if ch.isalnum())


def pick_column(headers: Sequence[str], candidates: Sequence[str]) -> Optional[str]:
    normalized = {normalize_header(header): header for header in headers}
    for candidate in candidates:
        key = normalize_header(candidate)
        if key in normalized:
            return normalized[key]
    for header in headers:
        header_norm = normalize_header(header)
        for candidate in candidates:
            if normalize_header(candidate) in header_norm:
                return header
    return None


def list_zip_members(zip_path: Path, allowed_ext: Iterable[str]) -> List[str]:
    if not zip_path.exists():
        return []
    with ZipFile(zip_path, "r") as handle:
        members = [name for name in handle.namelist() if name.lower().endswith(tuple(allowed_ext))]
    return members


def read_csv_from_zip(zip_path: Path, member_name: str) -> List[Dict[str, str]]:
    with ZipFile(zip_path, "r") as handle:
        with handle.open(member_name) as raw:
            text = TextIOWrapper(raw, encoding="utf-8", newline="")
            reader = csv.DictReader(text)
            return [row for row in reader]


def read_excel_from_zip(
    zip_path: Path,
    member_name: str,
    sheet_name: Optional[str] = None,
    max_rows: Optional[int] = None,
) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required to read Excel files. Install via requirements.txt") from exc
    with ZipFile(zip_path, "r") as handle:
        data = handle.read(member_name)
    workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in headers_row]
    output: List[Dict[str, Any]] = []
    for row in rows_iter:
        record = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        output.append(record)
        if max_rows is not None and len(output) >= max_rows:
            break
    return output


def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, sort_keys=False)


def parse_float(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None
    value = str(value).strip()
    if value == "" or value.lower() in {"na", "null"}:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def parse_int(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None
    value = str(value).strip()
    if value == "" or value.lower() in {"na", "null"}:
        return None
    try:
        return int(float(value))
    except ValueError:
        return None


def mean(values: List[Optional[float]]) -> Optional[float]:
    clean = [v for v in values if v is not None]
    if not clean:
        return None
    return sum(clean) / len(clean)
